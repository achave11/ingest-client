import re
import openpyxl

from ingest.importer.conversion import template_manager
from ingest.importer.conversion.template_manager import TemplateManager
from ingest.importer.spreadsheet.ingest_workbook import IngestWorkbook
from ingest.importer.submission import IngestSubmitter

from ingest.api.ingestapi import IngestApi

PROJECT_ID = 'dummy-project-id'


class IngestImporter:

    def __init__(self, ingest_api):
        self.ingest_api = ingest_api

    def import_spreadsheet(self, file_path, submission_url, dry_run=False):
        workbook = openpyxl.load_workbook(filename=file_path)
        ingest_workbook = IngestWorkbook(workbook)
        schemas = ingest_workbook.get_schemas()

        template_mgr = template_manager.build(schemas)
        workbook_importer = WorkbookImporter(template_mgr)
        spreadsheet_json = workbook_importer.do_import(ingest_workbook)

        submission = None
        if not dry_run:
            submitter = IngestSubmitter(self.ingest_api, template_mgr)
            submission = submitter.submit(spreadsheet_json, submission_url)
            print(f'Submission in {submission_url} is done!')

        return submission


class WorkbookImporter:

    def __init__(self, template_mgr):
        self.worksheet_importer = WorksheetImporter()
        self.template_mgr = template_mgr

    def do_import(self, workbook: IngestWorkbook):
        pre_ingest_json_map = {}

        self.import_project_worksheet(pre_ingest_json_map, workbook)

        for worksheet in workbook.importable_worksheets():
            entities_dict = self.worksheet_importer.do_import(worksheet, self.template_mgr)
            concrete_entity = self.template_mgr.get_concrete_entity_of_tab(worksheet.title)

            # TODO what if the tab is not a valid entity?
            if concrete_entity is None:
                print(f'{worksheet.title} is not a valid tab name.')
                continue

            domain_entity = self.template_mgr.get_domain_entity(concrete_entity)

            if pre_ingest_json_map.get(domain_entity) is None:
                pre_ingest_json_map[domain_entity] = {}

            pre_ingest_json_map[domain_entity].update(entities_dict)

        return pre_ingest_json_map

    def import_project_worksheet(self, pre_ingest_json_map, workbook):
        project_worksheet = workbook.get_project_worksheet()
        if project_worksheet:
            project_importer = ProjectWorksheetImporter()
            project = project_importer.do_import(project_worksheet, self.template_mgr)

            # TODO make this generic, determine which is the parent entity/worksheet then update it with the subtab contents
            contacts_worksheet = workbook.get_contacts_worksheet()

            if contacts_worksheet:
                contacts_importer = SubWorksheetImporter()
                contacts = contacts_importer.do_import(contacts_worksheet, self.template_mgr)
                project['contributors'] = contacts

            # TODO project has no identifier, put a dummy id for now to make the structure parallel with other entities
            pre_ingest_json_map['project'] = {}
            pre_ingest_json_map['project'][PROJECT_ID] = {}
            pre_ingest_json_map['project'][PROJECT_ID]['content'] = project


class WorksheetImporter:
    KEY_HEADER_ROW_IDX = 4
    USER_FRIENDLY_HEADER_ROW_IDX = 2
    START_ROW_IDX = 5

    def __init__(self):
        pass

    # TODO Refactor
    def do_import(self, worksheet, template:TemplateManager):
        concrete_entity = template.get_concrete_entity_of_tab(worksheet.title)

        # TODO if tab is not recognized, should we ignore it?
        if concrete_entity is None:
            return {}

        rows_by_id = {}
        for row in self._get_data_rows(worksheet):
            node = template.create_template_node(worksheet)
            object_list_tracker = ObjectListTracker()
            row_id = None

            links_map = {}

            for cell in row:
                # TODO preprocess headers so that cells can be converted without having to always
                # check the header
                header_key = self._get_header_name(cell, worksheet)

                cell_value = cell.value

                # TODO log warnings, confirm what to do if header name is not in schema template
                if header_key is None or cell_value is None:
                    continue

                field_chain = self._get_field_chain(header_key)

                if template.is_multivalue_object_field(header_key):
                    object_list_tracker.track_value(field_chain, cell_value)
                    continue

                converter = template.get_converter(header_key)
                data = converter.convert(cell_value)

                cell_concrete_entity = self._get_concrete_entity(template, header_key)


                if template.is_identifier_field(header_key):
                    if concrete_entity == cell_concrete_entity:
                        row_id = data
                    else:  # this is a link column

                        link_domain_entity = template.get_domain_entity(concrete_entity=cell_concrete_entity)

                        if not links_map.get(link_domain_entity):
                            links_map[link_domain_entity] = []

                        links_map[link_domain_entity].append(data)

                        continue  # do not add in content json

                node[field_chain] = data

            object_list_fields = object_list_tracker.get_object_list_fields()
            for field_chain in object_list_fields:
                node[field_chain] = object_list_tracker.get_value_by_field(field_chain)

            if not row_id:
                title = worksheet.title
                row_index = row[0].row  # get first cell row index
                raise NoUniqueIdFoundError(title, row_index)

            rows_by_id[row_id] = {
                'content': node.as_dict(),
                'links_by_entity': links_map
            }

        return rows_by_id

    def _get_data_rows(self, worksheet):
        return worksheet.iter_rows(row_offset=self.START_ROW_IDX, max_row=(worksheet.max_row - self.START_ROW_IDX))

    # TODO do not use user friendly names for now
    def _get_header_name(self, cell, worksheet):
        header_name = self._get_key_header_name(cell, worksheet)
        return header_name

    def _get_user_friendly_header_name(self, cell, worksheet):
        cell_template = '%s' + str(self.USER_FRIENDLY_HEADER_ROW_IDX)
        header_coordinate = cell_template % (cell.column)
        header_name = worksheet[header_coordinate].value
        return header_name

    def _get_key_header_name(self, cell, worksheet):
        cell_template = '%s' + str(self.KEY_HEADER_ROW_IDX)
        header_coordinate = cell_template % (cell.column)
        header_name = worksheet[header_coordinate].value
        return header_name

    def _get_field_chain(self, header_name):
        match = re.search('(\w+\.){1}(?P<field_chain>.*)', header_name)
        field_chain = match.group('field_chain')
        return field_chain

    def _get_concrete_entity(self, template_manager, header_name):
        return template_manager.get_concrete_entity_of_column(header_name)


class SubWorksheetImporter(WorksheetImporter):
    # TODO assumes for now that each row is always a part of an object
    # and all rows are a list field of a concrete entity
    # make this more generic??? or put the logic in Object list tracker, think more about this
    def do_import(self, worksheet, template: TemplateManager):
        rows = []
        for row in self._get_data_rows(worksheet):
            row_obj = {}

            for cell in row:
                header_key = self._get_header_name(cell, worksheet)
                cell_value = cell.value

                # TODO log warnings, confirm what to do if header name is not in schema template
                if header_key is None or cell_value is None:
                    continue

                field_name = self._get_subfield(header_key)

                converter = template.get_converter(header_key)
                data = converter.convert(cell_value)

                row_obj[field_name] = data

            rows.append(row_obj)

        return rows

    def _get_subfield(self, field_chain):
        match = re.search('(.*)\.(?P<field>\w+)', field_chain)
        return match.group('field')


class ProjectWorksheetImporter(WorksheetImporter):

    def do_import(self, worksheet, template:TemplateManager):
        # TODO add validation that this should only be one row
        rows = []
        for row in self._get_data_rows(worksheet):
            node = template.create_template_node(worksheet)
            object_list_tracker = ObjectListTracker()

            for cell in row:
                header_key = self._get_header_name(cell, worksheet)

                cell_value = cell.value

                # TODO log warnings, confirm what to do if header name is not in schema template
                if header_key is None or cell_value is None:
                    continue

                field_chain = self._get_field_chain(header_key)

                if template.is_multivalue_object_field(header_key):
                    object_list_tracker.track_value(field_chain, cell_value)
                    continue

                converter = template.get_converter(header_key)
                data = converter.convert(cell_value)
                node[field_chain] = data

            object_list_fields = object_list_tracker.get_object_list_fields()
            for field_chain in object_list_fields:
                node[field_chain] = object_list_tracker.get_value_by_field(field_chain)

            rows.append(node.as_dict())

        if len(rows) > 1:
            raise MultipleProjectsFound()

        return rows[0] if rows else None


class ObjectListTracker(object):

    def __init__(self):
        self.object_list_values = {}

    def _get_field(self, field_chain):
        match = re.search('(?P<field_chain>.*)(\.\w+)', field_chain)
        return match.group('field_chain')

    def _get_subfield(self, field_chain):
        match = re.search('(.*)\.(?P<field>\w+)', field_chain)
        return match.group('field')

    def track_value(self, header_name, value):
        subfield = self._get_subfield(header_name)
        field = self._get_field(header_name)

        if not self.object_list_values.get(field):
            self.object_list_values[field] = {}

        self.object_list_values[field][subfield] = value

    def get_object_list_fields(self):
        return self.object_list_values.keys()

    def get_value_by_field(self, field):
        return [self.object_list_values[field]]


class NoUniqueIdFoundError(Exception):
    def __init__(self, worksheet_title, row_index):
        message = f'No unique id found for row {row_index} in {worksheet_title} worksheet tab'
        super(NoUniqueIdFoundError, self).__init__(message)

        self.worksheet_title = worksheet_title
        self.row_index = row_index


class MultipleProjectsFound(Exception):
    pass
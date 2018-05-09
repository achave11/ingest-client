from unittest import TestCase

from mock import MagicMock, patch
from openpyxl import Workbook

from ingest.importer.conversion.data_converter import (
    Converter, ListConverter, BooleanConverter, DataType
)
from ingest.importer.conversion.template_manager import TemplateManager
from ingest.importer.importer import WorksheetImporter, WorkbookImporter, IngestWorkbook
from ingest.template.schematemplate import SchemaTemplate


def _create_single_row_worksheet(worksheet_data:dict):
    workbook = Workbook()
    worksheet = workbook.create_sheet()

    for column, data in worksheet_data.items():
        key, value = data
        worksheet[f'{column}1'] = key
        worksheet[f'{column}4'] = value

    return worksheet


class WorkbookImporterTest(TestCase):

    def test_do_import(self):
        # given:
        workbook_importer = WorkbookImporter()

        # and:
        workbook = Workbook()
        ingest_workbook = IngestWorkbook(workbook)

        # and:
        project_worksheet = workbook.create_sheet('Project')
        cell_suspension_worksheet = workbook.create_sheet('Cell Suspension')
        ingest_workbook.importable_worksheets = MagicMock(return_value=[
            project_worksheet, cell_suspension_worksheet
        ])

        # and:
        projects = [
            {'short_name': 'project 1', 'description': 'first project'},
            {'short_name': 'project 2', 'description': 'second project'}
        ]

        # and:
        cell_suspensions = [
            {'biomaterial_id': 'cell_suspension_101', 'biomaterial_name': 'cell suspension'}
        ]

        # and:
        worksheet_importer = WorksheetImporter()
        worksheet_importer.do_import = MagicMock(side_effect=[projects, cell_suspensions])

        # when:
        with patch('ingest.importer.importer.WorksheetImporter') as mock_constructor:
            mock_constructor.return_value = worksheet_importer
            pre_ingest_json_list = workbook_importer.do_import(ingest_workbook)

        # then:
        self.assertEqual(3, len(pre_ingest_json_list))
        for project in projects:
            self.assertTrue(project in pre_ingest_json_list, f'{project} not in list')
        for cell_suspension in cell_suspensions:
            self.assertTrue(cell_suspension in pre_ingest_json_list,
                            f'{cell_suspension} not in list')


class WorksheetImporterTest(TestCase):

    def test_do_import(self):
        # given:
        worksheet_importer = WorksheetImporter()

        # and:
        boolean_converter = BooleanConverter()
        converter_mapping = {
            'project.project_core.project_shortname': Converter(),
            'project.miscellaneous': ListConverter(),
            'project.numbers': ListConverter(data_type=DataType.INTEGER),
            'project.is_active': boolean_converter,
            'project.is_submitted': boolean_converter
        }

        # and:
        template_manager = TemplateManager(SchemaTemplate())
        template_manager.get_converter = lambda key: converter_mapping.get(key, Converter())

        # and:
        worksheet = self._create_test_worksheet()

        # when:
        json_list = worksheet_importer.do_import(worksheet, template_manager)

        # then:
        self.assertEqual(2, len(json_list))
        json = json_list[0]

        # and:
        self.assertTrue(2, len(json_list))
        self.assertEqual('Tissue stability 2', json_list[1]['project_core']['project_shortname'])

        project_core = json['project_core']
        self.assertEqual('Tissue stability', project_core['project_shortname'])
        self.assertEqual('Ischaemic sensitivity of human tissue by single cell RNA seq.',
                         project_core['project_title'])

        # and:
        self.assertEqual(2, len(json['miscellaneous']))
        self.assertEqual(['extra', 'details'], json['miscellaneous'])

        # and:
        self.assertEqual(7, json['contributor_count'])

        # and:
        self.assertEqual('Juan Dela Cruz||John Doe', json['contributors'])

        # and:
        self.assertEqual([1, 2, 3], json['numbers'])

        # and:
        self.assertEqual(True, json['is_active'])
        self.assertEqual(False, json['is_submitted'])

    def _create_test_worksheet(self):
        workbook = Workbook()
        worksheet = workbook.create_sheet('Project')
        worksheet['A1'] = 'project.project_core.project_shortname'
        worksheet['A4'] = 'Tissue stability'
        worksheet['A5'] = 'Tissue stability 2'
        worksheet['B1'] = 'project.project_core.project_title'
        worksheet['B4'] = 'Ischaemic sensitivity of human tissue by single cell RNA seq.'
        worksheet['C1'] = 'project.miscellaneous'
        worksheet['C4'] = 'extra||details'
        worksheet['D1'] = 'project.contributor_count'
        worksheet['D4'] = 7
        worksheet['E1'] = 'project.contributors'
        worksheet['E4'] = 'Juan Dela Cruz||John Doe'
        worksheet['F1'] = 'project.numbers'
        worksheet['F4'] = '1||2||3'
        worksheet['G1'] = 'project.is_active'
        worksheet['G4'] = 'Yes'
        worksheet['H1'] = 'project.is_submitted'
        worksheet['H4'] = 'No'

        return worksheet

    def test_do_import_with_ontology_fields(self):
        # given:
        template_manager = TemplateManager(SchemaTemplate())
        template_manager.get_converter = MagicMock(return_value=Converter())

        # and:
        ontology_fields_mapping = {
            'project.genus_species.ontology': True,
            'project.genus_species.text': True,
        }
        template_manager.is_ontology_subfield = (
            lambda field_name: ontology_fields_mapping.get(field_name)
        )

        # and:
        worksheet = _create_single_row_worksheet({
            'A': ('project.genus_species.ontology', 'UO:000008'),
            'B': ('project.genus_species.text', 'meter')
        })

        # and:
        worksheet_importer = WorksheetImporter()

        # when:
        json_list = worksheet_importer.do_import(worksheet, template_manager)

        # then:
        self.assertEqual(1, len(json_list))
        json = json_list[0]

        # and:
        self.assertTrue(type(json['genus_species']) is list)
        self.assertEqual(1, len(json['genus_species']))
        self.assertEqual({'ontology': 'UO:000008', 'text': 'meter'}, json['genus_species'][0])

    def test_do_import_adds_metadata_info(self):
        # given:
        template_manager = TemplateManager(SchemaTemplate())
        template_manager.get_converter = MagicMock(return_value=Converter())

        # and:
        # TODO merge get_schema_url with get_schema_type as predefined block of data
        # TODO the resulting method should be able to determine the data block based on
        #   the worksheet info (worksheet name)
        template_manager.get_schema_url = (
            lambda: 'https://schema.humancellatlas.org/type/project/5.1.0/project'
        )

        template_manager.get_schema_type = (
            lambda: 'project'
        )

        # and:
        importer = WorksheetImporter()

        # and:
        worksheet = _create_single_row_worksheet({
            'A': ('project.short_name', 'Project'),
            'B': ('project.description', 'This is a project')
        })

        # when:
        json_list = importer.do_import(worksheet, template_manager)

        # then:
        self.assertEqual(1, len(json_list))
        json = json_list[0]

        # and:
        self.assertEqual('https://schema.humancellatlas.org/type/project/5.1.0/project',
                         json['describedBy'])
        self.assertEqual('project', json['schema_type'])
